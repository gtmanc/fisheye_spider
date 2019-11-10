from openpyxl import load_workbook
from openpyxl import Workbook

import Constant

"""
Write cells in first row, in default sheet with specified names.
list: names will be written in the cells in first row

"""
def init(list):

    wb = Workbook()
    #wb = load_workbook('mainbuilding33.xlsx')
    ws = wb.active #get a defult sheet
    #print(ws.title)
    # ws.title = "first" #give a name to a sheet
    # ws2 = wb.create_sheet("my_sheet") #create a new sheet
    # ws = wb["first"] #get a heet with a specified name
    
    for i in range(len(list)):
        ws.cell(row = 1, column = i+1, value = list[i])

    return wb

"""
Write cells in a row starting with first cell. 
list: values to be wriiten to the the cell in the row
ws: worksheet
row_num: row number in the worksheet whichwill be written
"""
def write2ws(list, ws, row_num):
    if row_num == 0:
        print('invalid row number')
        return

    for i in range(len(list)):
        ws.cell(row = row_num, column = i+1, value = list[i])

    return row_num
